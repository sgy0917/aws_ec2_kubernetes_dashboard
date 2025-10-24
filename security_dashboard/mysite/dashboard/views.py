from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.db.models import Count, Q
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta
from collections import defaultdict
import json
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models import CheckRound, Asset, SecurityCheck


def check_api_server_status():
    """API 서버 상태 확인"""
    try:
        response = requests.get(f"{settings.API_SERVER_URL}/health", timeout=5)
        if response.status_code == 200:
            return 'running'
    except:
        pass
    return 'stopped'


def index(request):
    """
    메인 대시보드
    - 회차별 점검 통계 그래프
    - 자산 비용 그래프
    - 점검 시작 버튼
    """
    # 최근 7일간의 회차별 통계
    seven_days_ago = datetime.now().date() - timedelta(days=7)
    recent_rounds = CheckRound.objects.filter(
        check_date__gte=seven_days_ago
    ).order_by('check_date', 'round_number')
    
    # 그래프 데이터 생성
    chart_dates = []
    chart_passed = []
    chart_failed = []
    chart_warnings = []
    
    for round_obj in recent_rounds:
        stats = round_obj.get_statistics()
        chart_dates.append(f"{round_obj.check_date.strftime('%m/%d')} {round_obj.round_number}회차")
        chart_passed.append(stats['total_passed'])
        chart_failed.append(stats['total_failed'])
        chart_warnings.append(stats['total_warnings'])
    
    # 자산별 비용 계산 (전체 자산 기준)
    assets = Asset.objects.all()
    asset_costs = []
    
    for asset in assets:
        distro = asset.distro.lower() if asset.distro else ''
        if 'alpine' in distro:
            vcpu = 1
            memory_gb = 1
            cost = (vcpu * 50) + (memory_gb * 10)
        elif 'ubuntu' in distro:
            vcpu = 2
            memory_gb = 4
            cost = (vcpu * 50) + (memory_gb * 10)
        elif 'debian' in distro:
            vcpu = 2
            memory_gb = 2
            cost = (vcpu * 50) + (memory_gb * 10)
        else:
            vcpu = 1
            memory_gb = 1
            cost = (vcpu * 50) + (memory_gb * 10)
        
        asset_costs.append({
            'name': asset.name,
            'vcpu': vcpu,
            'memory_gb': memory_gb,
            'cost': cost
        })
    
    # 전체 통계
    total_rounds = CheckRound.objects.count()
    latest_round = CheckRound.objects.order_by('-check_date', '-round_number').first()
    
    # 최근 점검 회차의 자산 개수 (카드 표시용)
    latest_round_assets = 0
    if latest_round:
        latest_round_assets = latest_round.get_total_assets()
    
    latest_stats = {}
    if latest_round:
        latest_stats = latest_round.get_statistics()
    
    # 인프라 월간 비용 (전체 자산 기준)
    infrastructure_cost = sum(item['cost'] for item in asset_costs) if asset_costs else 0
    
    # 점검 비용 계산
    CHECK_COST_PER_RUN = 10
    total_check_cost = total_rounds * CHECK_COST_PER_RUN
    
    # 총 비용
    total_cost = infrastructure_cost + total_check_cost
    
    # API 서버 상태 확인
    api_server_status = check_api_server_status()
    
    context = {
        'latest_round_assets': latest_round_assets,  # 최근 점검한 자산 개수
        'total_rounds': total_rounds,
        'latest_round': latest_round,
        'latest_stats': latest_stats,
        'total_cost': total_cost,
        'infrastructure_cost': infrastructure_cost,
        'total_check_cost': total_check_cost,
        'check_cost_per_run': CHECK_COST_PER_RUN,
        'api_server_status': api_server_status,
        
        'chart_dates': json.dumps(chart_dates),
        'chart_passed': json.dumps(chart_passed),
        'chart_failed': json.dumps(chart_failed),
        'chart_warnings': json.dumps(chart_warnings),
        
        'asset_costs': asset_costs,
        'cost_labels': json.dumps([item['name'] for item in asset_costs]),
        'cost_data': json.dumps([item['cost'] for item in asset_costs]),
    }
    
    return render(request, 'dashboard/index.html', context)


def rounds_list(request):
    """회차 목록 페이지"""
    filter_date = request.GET.get('date', '')
    filter_round = request.GET.get('round', '')
    
    # 디버깅용 출력
    print(f"[필터링] 받은 날짜: '{filter_date}', 받은 회차: '{filter_round}'")
    
    rounds_query = CheckRound.objects.all()
    
    # 날짜 필터
    if filter_date:
        try:
            # 여러 날짜 형식 시도
            for date_format in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']:
                try:
                    filter_date_obj = datetime.strptime(filter_date, date_format).date()
                    rounds_query = rounds_query.filter(check_date=filter_date_obj)
                    print(f"[필터링] 날짜 필터 적용: {filter_date_obj}")
                    break
                except ValueError:
                    continue
        except Exception as e:
            print(f"[필터링] 날짜 파싱 실패: {e}")
    
    # 회차 필터
    if filter_round:
        try:
            filter_round_num = int(filter_round)
            rounds_query = rounds_query.filter(round_number=filter_round_num)
            print(f"[필터링] 회차 필터 적용: {filter_round_num}")
        except ValueError as e:
            print(f"[필터링] 회차 파싱 실패: {e}")
    
    # 필터 결과 확인
    print(f"[필터링] 결과 개수: {rounds_query.count()}개")
    
    rounds_by_date = defaultdict(list)
    
    for round_obj in rounds_query:
        round_data = {
            'id': round_obj.id,
            'round_number': round_obj.round_number,
            'check_time': round_obj.check_time.strftime('%H:%M:%S'),
            'datetime_str': round_obj.get_datetime_str(),
            'total_assets': round_obj.get_total_assets(),
            'stats': round_obj.get_statistics(),
        }
        rounds_by_date[round_obj.check_date].append(round_data)
    
    rounds_list = []
    for date in sorted(rounds_by_date.keys(), reverse=True):
        rounds_list.append({
            'date': date,
            'rounds': rounds_by_date[date]
        })
    
    # 전체 날짜 목록
    all_dates = CheckRound.objects.values_list('check_date', flat=True).distinct().order_by('-check_date')
    
    # 전체 회차 목록 (중복 제거하고 정렬)
    all_rounds = CheckRound.objects.values_list('round_number', flat=True).distinct().order_by('round_number')
    
    context = {
        'rounds_list': rounds_list,
        'all_dates': all_dates,
        'all_rounds': all_rounds,
        'filter_date': filter_date,
        'filter_round': filter_round,
    }
    
    return render(request, 'dashboard/rounds_list.html', context)


def round_detail(request, round_id):
    """특정 회차의 자산 목록"""
    check_round = get_object_or_404(CheckRound, id=round_id)
    
    # 이 회차의 보안 점검 결과들
    security_checks = SecurityCheck.objects.filter(round=check_round).select_related('asset')
    
    # 자산 정보 구성
    assets = []
    for sec_check in security_checks:
        asset_data = {
            'id': sec_check.asset.id,
            'asset_code': sec_check.asset.asset_code,
            'name': sec_check.asset.name,
            'hostname': sec_check.asset.hostname,
            'distro': sec_check.asset.distro,
            'os_version': sec_check.asset.os_version,
            'execution_type': sec_check.asset.execution_type,
            'check_date': sec_check.check_date,
            'passed': sec_check.passed_checks,
            'warnings': sec_check.warning_checks,
            'failed': sec_check.failed_checks,
            'na': sec_check.not_applicable_checks,
        }
        assets.append(asset_data)
    
    context = {
        'check_round': check_round,
        'assets': assets,
        'total_assets': len(assets),
    }
    
    return render(request, 'dashboard/list.html', context)


def asset_detail(request, asset_id):
    """자산 상세 페이지"""
    asset = get_object_or_404(Asset, id=asset_id)
    
    # 이 자산의 최근 보안 점검 결과
    latest_security_check = SecurityCheck.objects.filter(asset=asset).order_by('-check_date').first()
    
    if not latest_security_check:
        return render(request, 'dashboard/asset_detail.html', {
            'asset': asset,
            'error': '점검 결과가 없습니다.'
        })
    
    check_round = latest_security_check.round
    
    # 상태 필터
    status_filter = request.GET.get('status', '')
    
    # details JSON에서 점검 항목 추출
    check_details_raw = latest_security_check.details if latest_security_check.details else []
    
    # 필터링
    if status_filter:
        filtered_details = []
        for check_item in check_details_raw:
            item_status = check_item.get('status', '').lower()
            if status_filter == 'pass' and item_status in ['pass', 'passed', '양호']:
                filtered_details.append(check_item)
            elif status_filter == 'fail' and item_status in ['fail', 'failed', '취약']:
                filtered_details.append(check_item)
            elif status_filter == 'warn' and item_status in ['warn', 'warning', '주의']:
                filtered_details.append(check_item)
            elif status_filter == 'not_applicable' and item_status in ['not_applicable', 'n/a', 'na', '해당없음']:
                filtered_details.append(check_item)
        check_details_raw = filtered_details
    
    # 통계
    total_checks = latest_security_check.total_checks
    passed_checks = latest_security_check.passed_checks
    failed_checks = latest_security_check.failed_checks
    warning_checks = latest_security_check.warning_checks
    not_applicable_checks = latest_security_check.not_applicable_checks
    
    pass_rate = latest_security_check.get_pass_rate()
    
    # 차트 데이터
    chart_labels = ['양호', '주의', '취약', '해당없음']
    chart_data = [passed_checks, warning_checks, failed_checks, not_applicable_checks]
    
    # 점검 항목 상세 정보 구성
    check_details = []
    for idx, check_item in enumerate(check_details_raw):
        status = check_item.get('status', 'unknown').lower()
        
        status_normalized = 'pass'
        status_display = '양호'
        
        if status in ['fail', 'failed', '취약']:
            status_normalized = 'fail'
            status_display = '취약'
        elif status in ['warn', 'warning', '주의']:
            status_normalized = 'warn'
            status_display = '주의'
        elif status in ['not_applicable', 'n/a', 'na', '해당없음']:
            status_normalized = 'not_applicable'
            status_display = '해당없음'
        
        check_details.append({
            'id': check_item.get('check_id', f'U-{idx+1:02d}'),
            'name': check_item.get('name', check_item.get('check_name', '점검 항목')),
            'status': check_item.get('status', 'unknown'),
            'status_normalized': status_normalized,
            'status_display': status_display,
            'details': check_item.get('details', ''),
            'checked_paths': check_item.get('checked_paths', ''),
            'commands_executed': check_item.get('commands_executed', ''),
            'recommendation': check_item.get('recommendation', ''),
        })
    
    context = {
        'asset': asset,
        'check_round': check_round,
        'latest_check': latest_security_check,
        'total_checks': total_checks,
        'passed_checks': passed_checks,
        'failed_checks': failed_checks,
        'warning_checks': warning_checks,
        'not_applicable_checks': not_applicable_checks,
        'pass_rate': pass_rate,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        'check_details': check_details,
        'status_filter': status_filter,
    }
    
    return render(request, 'dashboard/asset_detail.html', context)


@csrf_exempt
def start_check(request):
    """보안 점검 시작 API (CSRF 제외)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST 요청만 허용됩니다.'}, status=405)
    
    try:
        # API 서버에 점검 시작 요청
        api_url = f"{settings.API_SERVER_URL}/start_check"
        response = requests.post(api_url, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            return JsonResponse({
                'success': True,
                'message': '보안 점검이 시작되었습니다.',
                'data': data
            })
        else:
            return JsonResponse({
                'success': False,
                'message': f'API 서버 오류: {response.status_code}'
            }, status=500)
            
    except requests.exceptions.RequestException as e:
        return JsonResponse({
            'success': False,
            'message': f'API 서버 연결 실패: {str(e)}'
        }, status=500)


def export_asset_excel(request, asset_id):
    """자산 상세 정보 엑셀 다운로드"""
    asset = get_object_or_404(Asset, id=asset_id)
    
    # 이 자산의 최근 보안 점검 결과
    latest_security_check = SecurityCheck.objects.filter(asset=asset).order_by('-check_date').first()
    
    if not latest_security_check:
        return HttpResponse("점검 결과가 없습니다.", status=404)
    
    check_details = latest_security_check.details if latest_security_check.details else []
    
    # 엑셀 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{asset.asset_code}"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # 자산 정보 헤더
    ws['A1'] = '자산 정보'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    # 자산 정보
    ws['A2'] = '자산 코드'
    ws['B2'] = asset.asset_code
    ws['A3'] = '호스트'
    ws['B3'] = asset.hostname or 'N/A'
    ws['A4'] = '배포판'
    ws['B4'] = asset.distro or 'Unknown'
    ws['A5'] = 'OS 버전'
    ws['B5'] = asset.os_version or 'Unknown'
    ws['A6'] = '커널'
    ws['B6'] = asset.kernel or 'N/A'
    ws['A7'] = '점검 일시'
    ws['B7'] = latest_security_check.check_date.strftime('%Y-%m-%d %H:%M:%S')
    
    # 점검 항목 헤더
    ws['A9'] = '점검 항목'
    ws['A9'].font = Font(bold=True, size=14)
    ws.merge_cells('A9:F9')
    
    # 테이블 헤더
    headers = ['점검 ID', '점검명', '상태', '점검 내용', '조치사항', '점검 경로']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 점검 데이터
    row = 11
    for idx, check_item in enumerate(check_details):
        ws.cell(row=row, column=1, value=check_item.get('check_id', f'U-{idx+1:02d}'))
        ws.cell(row=row, column=2, value=check_item.get('name', check_item.get('check_name', '')))
        ws.cell(row=row, column=3, value=check_item.get('status', ''))
        ws.cell(row=row, column=4, value=check_item.get('details', ''))
        ws.cell(row=row, column=5, value=check_item.get('recommendation', ''))
        ws.cell(row=row, column=6, value=check_item.get('checked_paths', ''))
        row += 1
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 40
    
    # HTTP 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={asset.asset_code}_check_result.xlsx'
    
    wb.save(response)
    return response
