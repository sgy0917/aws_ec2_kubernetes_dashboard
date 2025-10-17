from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.db.models import Count, Q
from datetime import datetime, timedelta
from collections import defaultdict
from dashboard.models import Asset, SecurityCheck, CheckRound
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def dashboard(request):
    """
    메인 대시보드 - 전부 내림차순 정렬
    """
    filter_date = request.GET.get('date', '')
    filter_round = request.GET.get('round', '')

    # 전부 내림차순: 날짜 DESC, 시간 DESC, 회차번호 DESC
    rounds_query = CheckRound.objects.all().order_by('-check_date', '-check_time', '-round_number')

    # 날짜 필터링
    if filter_date:
        try:
            filter_date_obj = datetime.strptime(filter_date, '%Y-%m-%d').date()
            rounds_query = rounds_query.filter(check_date=filter_date_obj)
        except ValueError:
            pass

    # 회차 필터링
    if filter_round:
        try:
            filter_round_num = int(filter_round)
            rounds_query = rounds_query.filter(round_number=filter_round_num)
        except ValueError:
            pass

    # 날짜별로 그룹화 (이미 정렬되어 있음)
    rounds_by_date = defaultdict(list)

    for round_obj in rounds_query:
        round_data = {
            'id': round_obj.id,
            'round_number': round_obj.round_number,
            'check_time': round_obj.check_time.strftime('%H:%M:%S'),
            'datetime_str': round_obj.get_datetime_str(),
            'total_assets': round_obj.get_total_assets(),
            'stats': round_obj.get_statistics(),
        }
        rounds_by_date[round_obj.check_date].append(round_data)

    # 날짜 내림차순으로 리스트 생성
    rounds_list = []
    for date in sorted(rounds_by_date.keys(), reverse=True):
        rounds_list.append({
            'date': date,
            'rounds': rounds_by_date[date]  # 이미 내림차순 정렬됨
        })

    # 필터 옵션
    all_dates = CheckRound.objects.values_list('check_date', flat=True).distinct().order_by('-check_date')
    all_rounds = CheckRound.objects.values_list('round_number', flat=True).distinct().order_by('-round_number')

    context = {
        'rounds_list': rounds_list,
        'all_dates': all_dates,
        'all_rounds': all_rounds,
        'filter_date': filter_date,
        'filter_round': filter_round,
    }

    return render(request, 'dashboard/index.html', context)


def round_detail(request, round_id):
    """
    회차별 자산 목록
    """
    check_round = get_object_or_404(CheckRound, id=round_id)

    # 해당 회차의 모든 점검 결과
    security_checks = SecurityCheck.objects.filter(round=check_round).select_related('asset')

    # 자산 목록 생성
    assets_data = []
    for check in security_checks:
        asset = check.asset
        assets_data.append({
            'id': asset.id,
            'asset_code': asset.asset_code,
            'name': asset.name,
            'hostname': asset.hostname,
            'distro': asset.distro,
            'os_version': asset.os_version,
            'execution_type': asset.execution_type,
            'check_date': check.check_date,
            'passed': check.passed_checks,
            'warnings': check.warning_checks,
            'failed': check.failed_checks,
            'na': check.not_applicable_checks,
            'status': check.status,
        })

    context = {
        'check_round': check_round,
        'assets': assets_data,
        'total_assets': len(assets_data),
    }

    return render(request, 'dashboard/list.html', context)


def asset_detail(request, asset_id):
    """
    자산 상세 페이지
    """
    asset = get_object_or_404(Asset, id=asset_id)

    # 최근 점검 결과
    latest_check = SecurityCheck.objects.filter(asset=asset).order_by('-check_date').first()

    if not latest_check:
        context = {
            'asset': asset,
            'error_message': '점검 데이터가 없습니다.'
        }
        return render(request, 'dashboard/asset_detail.html', context)

    # 상세 점검 항목
    check_details = latest_check.details if latest_check.details else []

    # 상태별 필터링
    status_filter = request.GET.get('status', '')
    if status_filter:
        if status_filter == 'pass':
            check_details = [c for c in check_details if c.get('status') in ['양호', 'pass']]
        elif status_filter == 'fail':
            check_details = [c for c in check_details if c.get('status') in ['취약', 'fail']]
        elif status_filter == 'warn':
            check_details = [c for c in check_details if c.get('status') in ['주의', 'warn']]
        elif status_filter == 'not_applicable':
            check_details = [c for c in check_details if c.get('status') in ['해당없음', 'not_applicable', '해당 없음']]

    # 상태 정규화
    for check in check_details:
        status = check.get('status', '')
        if status in ['pass', 'PASS', '양호']:
            check['status_normalized'] = 'pass'
            check['status_display'] = '양호'
        elif status in ['fail', 'FAIL', '취약']:
            check['status_normalized'] = 'fail'
            check['status_display'] = '취약'
        elif status in ['warn', 'WARN', '주의']:
            check['status_normalized'] = 'warn'
            check['status_display'] = '주의'
        else:
            check['status_normalized'] = 'not_applicable'
            check['status_display'] = '해당없음'

    # Chart.js용 데이터
    chart_labels = json.dumps(['양호', '주의', '취약', '해당없음'])
    chart_data = json.dumps([
        latest_check.passed_checks,
        latest_check.warning_checks,
        latest_check.failed_checks,
        latest_check.not_applicable_checks
    ])

    context = {
        'asset': asset,
        'latest_check': latest_check,
        'check_round': latest_check.round,
        'check_details': check_details,
        'total_checks': latest_check.total_checks,
        'passed_checks': latest_check.passed_checks,
        'warning_checks': latest_check.warning_checks,
        'failed_checks': latest_check.failed_checks,
        'not_applicable_checks': latest_check.not_applicable_checks,
        'pass_rate': latest_check.get_pass_rate(),
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'status_filter': status_filter,
    }

    return render(request, 'dashboard/asset_detail.html', context)


def export_asset_excel(request, asset_id):
    """자산 점검 결과 Excel 다운로드"""
    asset = get_object_or_404(Asset, id=asset_id)
    latest_check = SecurityCheck.objects.filter(asset=asset).order_by('-check_date').first()

    if not latest_check:
        return HttpResponse("점검 데이터가 없습니다.", status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "점검 결과"

    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    na_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 제목
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"보안 점검 결과 리포트 - {asset.name}"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 자산 정보
    ws['A3'] = "자산 정보"
    ws['A3'].font = Font(bold=True, size=14)

    info_data = [
        ["자산 코드", asset.asset_code],
        ["자산명", asset.name],
        ["호스트", asset.hostname or "N/A"],
        ["배포판", asset.distro or "Unknown"],
        ["OS 버전", asset.os_version or "Unknown"],
        ["커널 버전", asset.kernel or "N/A"],
        ["실행 방식", asset.execution_type or "N/A"],
        ["점검 일시", latest_check.check_date.strftime("%Y-%m-%d %H:%M:%S")],
    ]

    row = 4
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    # 통계
    row += 1
    ws[f'A{row}'] = "점검 통계"
    ws[f'A{row}'].font = Font(bold=True, size=14)

    row += 1
    stats_data = [
        ["총 점검 수", latest_check.total_checks],
        ["양호", latest_check.passed_checks],
        ["주의", latest_check.warning_checks],
        ["취약", latest_check.failed_checks],
        ["해당없음", latest_check.not_applicable_checks],
        ["합격률", f"{latest_check.get_pass_rate()}%"],
    ]

    for label, value in stats_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    # 상세 점검 결과
    row += 2
    ws[f'A{row}'] = "상세 점검 결과"
    ws[f'A{row}'].font = Font(bold=True, size=14)

    row += 1
    headers = ["번호", "점검ID", "점검명", "상태", "점검내용", "점검경로", "실행명령어", "권장사항"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 컬럼 너비
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 40

    # 상세 데이터
    if latest_check.details:
        for idx, check in enumerate(latest_check.details, 1):
            row += 1

            status = check.get('status', '')

            if status in ['양호', 'pass']:
                status_kr = '양호'
                fill = pass_fill
            elif status in ['취약', 'fail']:
                status_kr = '취약'
                fill = fail_fill
            elif status in ['주의', 'warn']:
                status_kr = '주의'
                fill = warn_fill
            else:
                status_kr = '해당없음'
                fill = na_fill

            data_row = [
                idx,
                check.get('id', ''),
                check.get('name', ''),
                status_kr,
                check.get('details', ''),
                check.get('checked_paths', ''),
                check.get('commands_executed', ''),
                check.get('recommendation', '')
            ]

            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

                if col == 4:
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    # HTTP 응답
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    filename = f"security_check_{asset.asset_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
